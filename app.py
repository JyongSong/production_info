from __future__ import annotations

from io import BytesIO
import os
from pathlib import Path

from flask import Flask, jsonify, render_template, request, send_file, redirect
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import supabase_client as db
from database import initialize_database
from services import (
    DuplicatePairError,
    DuplicateQRCodeError,
    DuplicateSaveError,
    InvalidLumiSnError,
    MatchServiceError,
    NotFoundError,
    ValidationError,
    check_lumi_sn_already_used,
    check_lumi_sn_exists,
    count_matches_by_date,
    create_match,
    delete_match,
    get_qr_settings,
    get_recent_matches,
    list_matches_for_export,
    search_matches,
    update_match,
    update_qr_settings,
)
from services.match_service import get_kst_now

LUMI_PRODUCT_TABLE = "lumi_product_sn"


app = Flask(__name__)
app.json.ensure_ascii = False

initialize_database()

BASE_DIR = Path(__file__).resolve().parent

# Read Access Password (empty means password authentication disabled)
ACCESS_PASSWORD = os.environ.get("ACCESS_PASSWORD", "").strip()


@app.before_request
def require_login():
    if not ACCESS_PASSWORD:
        return
    
    # Allow static assets and login endpoints
    if request.path.startswith("/static/") or request.path in ["/login", "/api/login"]:
        return
        
    token = request.cookies.get("access_token")
    if token != ACCESS_PASSWORD:
        return redirect("/login")


def build_dashboard_payload(message: str, match: dict | None = None) -> dict:
    today = get_kst_now().strftime("%Y-%m-%d")
    payload = {
        "success": True,
        "message": message,
        "recent_matches": get_recent_matches(),
        "today_count": count_matches_by_date(today),
    }
    if match is not None:
        payload["match"] = match
    return payload


def get_asset_version(relative_path: str) -> str:
    asset_path = BASE_DIR / relative_path
    if not asset_path.exists():
        return "1"
    return str(int(asset_path.stat().st_mtime))


def build_excel_file(rows: list[dict]) -> BytesIO:
    # 현장 PC에서 바로 열어도 보기 쉽도록 헤더와 폭을 함께 정리한다.
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "QR 매칭 내역"

    headers = ["번호", "Lumi SN", "Solity SN", "Production Time", "작업자명", "비고"]
    worksheet.append(headers)

    for row in rows:
        worksheet.append(
            [
                row["id"],
                row["first_qr"],
                row["second_qr"],
                row["created_at"],
                row["operator_name"],
                row["note"],
            ]
        )

    header_fill = PatternFill(fill_type="solid", start_color="123C52", end_color="123C52")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    column_widths = [10, 32, 32, 22, 18, 24]
    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for row in worksheet.iter_rows(min_row=2, max_col=6):
        row[0].alignment = center_alignment
        row[3].alignment = center_alignment

    worksheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@app.after_request
def add_no_cache_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


def _common_versions() -> dict:
    return {
        "css_version": get_asset_version("static/css/style.css"),
        "js_version": get_asset_version("static/js/app.js"),
        "scan_js_version": get_asset_version("static/js/scan-core.js"),
        "sound_js_version": get_asset_version("static/js/sound.js"),
    }


@app.get("/login")
def login_page():
    return render_template("login.html")


@app.post("/api/login")
def login_api():
    payload = request.get_json(silent=True) or request.form
    password = (payload.get("password", "") or "").strip()
    
    if password == ACCESS_PASSWORD:
        resp = jsonify({"success": True, "message": "로그인에 성공했습니다."})
        resp.set_cookie("access_token", ACCESS_PASSWORD, max_age=30*24*60*60, httponly=True)
        return resp
    else:
        return jsonify({"success": False, "message": "비밀번호가 올바르지 않습니다."}), 401


@app.get("/")
def home():
    today = get_kst_now().strftime("%Y-%m-%d")
    return render_template(
        "home.html",
        today_count=count_matches_by_date(today),
        today=today,
        **_common_versions(),
    )


@app.get("/scan")
def scan_page():
    today = get_kst_now().strftime("%Y-%m-%d")
    return render_template(
        "scan.html",
        recent_matches=get_recent_matches(),
        today_count=count_matches_by_date(today),
        today=today,
        qr_settings=get_qr_settings(),
        **_common_versions(),
    )


@app.get("/search")
def search_page():
    return render_template("search.html", **_common_versions())


@app.get("/settings")
def settings_page():
    # Count existing Lumi SN entries
    try:
        rows = db.select(LUMI_PRODUCT_TABLE, columns="lumi_sn")
        lumi_sn_count = len(rows)
    except Exception:
        lumi_sn_count = 0

    return render_template(
        "settings.html",
        qr_settings=get_qr_settings(),
        lumi_sn_count=lumi_sn_count,
        **_common_versions(),
    )


def error_response(error: MatchServiceError, status: int):
    """Return a JSON error that tells the client which input to focus."""
    return jsonify({
        "success": False,
        "message": str(error),
        "field": getattr(error, "field", None),
    }), status


@app.post("/api/matches")
def save_match():
    payload = request.get_json(silent=True) or request.form

    try:
        # 스캐너 입력은 프런트엔드에서 JSON으로 보내고, 일반 폼 전송도 함께 허용한다.
        saved_match = create_match(
            first_qr=payload.get("first_qr", ""),
            second_qr=payload.get("second_qr", ""),
            operator_name=payload.get("operator_name", ""),
            note=payload.get("note", ""),
        )
        return jsonify(build_dashboard_payload("매칭이 저장되었습니다.", saved_match))
    except ValidationError as error:
        return error_response(error, 400)
    except InvalidLumiSnError as error:
        return error_response(error, 400)
    except DuplicatePairError as error:
        return error_response(error, 409)
    except DuplicateQRCodeError as error:
        return error_response(error, 409)
    except DuplicateSaveError as error:
        return error_response(error, 409)
    except Exception:
        return jsonify({"success": False, "message": "저장 중 오류가 발생했습니다."}), 500


@app.put("/api/matches/<int:match_id>")
def update_match_api(match_id: int):
    payload = request.get_json(silent=True) or request.form

    try:
        updated_match = update_match(
            match_id=match_id,
            first_qr=payload.get("first_qr", ""),
            second_qr=payload.get("second_qr", ""),
            operator_name=payload.get("operator_name", ""),
            note=payload.get("note", ""),
        )
        return jsonify(build_dashboard_payload("수정이 완료되었습니다.", updated_match))
    except NotFoundError as error:
        return jsonify({"success": False, "message": str(error)}), 404
    except ValidationError as error:
        return error_response(error, 400)
    except InvalidLumiSnError as error:
        return error_response(error, 400)
    except DuplicatePairError as error:
        return error_response(error, 409)
    except DuplicateQRCodeError as error:
        return error_response(error, 409)
    except DuplicateSaveError as error:
        return error_response(error, 409)
    except Exception:
        return jsonify({"success": False, "message": "수정 중 오류가 발생했습니다."}), 500


@app.delete("/api/matches/<int:match_id>")
def delete_match_api(match_id: int):
    try:
        delete_match(match_id)
        return jsonify(build_dashboard_payload("삭제가 완료되었습니다."))
    except NotFoundError as error:
        return jsonify({"success": False, "message": str(error)}), 404
    except Exception:
        return jsonify({"success": False, "message": "삭제 중 오류가 발생했습니다."}), 500


@app.get("/api/validate-lumi-sn")
def validate_lumi_sn_api():
    lumi_sn = (request.args.get("sn", "") or "").strip()
    if not lumi_sn:
        return jsonify({"valid": False, "message": "Lumi SN을 입력해주세요."}), 400
    if not check_lumi_sn_exists(lumi_sn):
        return jsonify({"valid": False, "message": "등록되지 않은 Lumi SN입니다."})
    if check_lumi_sn_already_used(lumi_sn):
        return jsonify({"valid": False, "message": "이미 사용된 Lumi SN입니다."})
    return jsonify({"valid": True})


@app.get("/api/settings")
def get_settings_api():
    return jsonify({"settings": get_qr_settings()})


@app.put("/api/settings")
def update_settings_api():
    payload = request.get_json(silent=True) or request.form

    try:
        settings = update_qr_settings(
            first_qr_length=payload.get("first_qr_length", 0),
            second_qr_length=payload.get("second_qr_length", 0),
        )
        return jsonify(
            {
                "success": True,
                "message": "QR 자릿수 설정이 저장되었습니다.",
                "settings": settings,
            }
        )
    except ValidationError as error:
        return jsonify({"success": False, "message": str(error)}), 400


@app.get("/api/recent")
def recent_matches_api():
    today = get_kst_now().strftime("%Y-%m-%d")
    return jsonify(
        {
            "recent_matches": get_recent_matches(),
            "today_count": count_matches_by_date(today),
        }
    )


@app.get("/api/search")
def search_matches_api():
    try:
        matches = search_matches(
            first_qr=request.args.get("first_qr", ""),
            second_qr=request.args.get("second_qr", ""),
            target_date=request.args.get("date", ""),
        )
        return jsonify({"matches": matches, "count": len(matches)})
    except ValidationError as error:
        return jsonify({"success": False, "message": str(error)}), 400


@app.post("/api/upload-lumi-sn")
def upload_lumi_sn():
    """Upload an Excel file, convert to CSV-like data, and replace lumi_product_sn table."""
    if "file" not in request.files:
        return jsonify({"success": False, "message": "파일이 첨부되지 않았습니다."}), 400

    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return jsonify({"success": False, "message": ".xlsx 파일만 업로드할 수 있습니다."}), 400

    try:
        from openpyxl import load_workbook

        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        # Read header row
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value or "").strip().lower())

        if "lumi_sn" not in headers:
            wb.close()
            return jsonify({"success": False, "message": "Excel 파일에 'lumi_sn' 열이 필요합니다."}), 400

        sn_col_idx = headers.index("lumi_sn")
        time_col_idx = headers.index("time") if "time" in headers else None

        # Read data rows
        rows_to_insert = []
        seen = set()
        for row in ws.iter_rows(min_row=2):
            sn_value = str(row[sn_col_idx].value or "").strip()
            if not sn_value or sn_value in seen:
                continue
            seen.add(sn_value)

            record = {"lumi_sn": sn_value}
            if time_col_idx is not None:
                time_val = row[time_col_idx].value
                if time_val is not None:
                    record["time"] = str(time_val).strip()
                else:
                    record["time"] = ""
            rows_to_insert.append(record)

        wb.close()

        if not rows_to_insert:
            return jsonify({"success": False, "message": "유효한 데이터가 없습니다."}), 400

        # Replace the entire whitelist atomically in a transaction
        db.rpc("replace_lumi_product_sn", {"p_rows": rows_to_insert})

        return jsonify({
            "success": True,
            "message": f"Lumi SN {len(rows_to_insert)}건이 업로드되었습니다.",
            "count": len(rows_to_insert),
        })

    except Exception as e:
        return jsonify({"success": False, "message": f"파일 처리 중 오류: {str(e)}"}), 500


@app.get("/api/lumi-sn-sample")
def download_lumi_sn_sample():
    """Download a sample Excel file based on lumi_product_sn table headers and first 10 rows."""
    try:
        rows = db.select(
            LUMI_PRODUCT_TABLE,
            columns="lumi_sn,time",
            limit=10,
        )
    except Exception:
        rows = []

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "lumi_product_sn"

    # Headers
    sample_headers = ["lumi_sn", "time"]
    worksheet.append(sample_headers)

    header_fill = PatternFill(fill_type="solid", start_color="123C52", end_color="123C52")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row in rows:
        worksheet.append([
            row.get("lumi_sn", ""),
            row.get("time", ""),
        ])

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 22
    worksheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="lumi_sn_sample.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/lumi-whitelist-status")
def get_whitelist_status():
    """Get whitelist progress stats: total, used, unused."""
    try:
        whitelist = db.select(LUMI_PRODUCT_TABLE, columns="lumi_sn")
        total_count = len(whitelist)

        # Get all non-deleted production matches
        matches = db.select("production_records", columns="lumi_sn", filters={"deleted_at": "is.null"})
        used_lumi_sns = {m["lumi_sn"] for m in matches}

        # Calculate overlap
        used_count = sum(1 for item in whitelist if item["lumi_sn"] in used_lumi_sns)
        unused_count = total_count - used_count

        return jsonify({
            "success": True,
            "total_count": total_count,
            "used_count": used_count,
            "unused_count": unused_count,
            "progress_percent": round((used_count / total_count * 100), 1) if total_count > 0 else 0
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.get("/download.xlsx")
def download_excel():
    try:
        target_date = request.args.get("date", "")
        rows = list_matches_for_export(target_date=target_date)
        excel_file = build_excel_file(rows)

        date_suffix = target_date.strip() if target_date else get_kst_now().strftime("%Y%m%d_%H%M%S")
        filename = f"qr_matching_{date_suffix}.xlsx"
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ValidationError as error:
        return jsonify({"success": False, "message": str(error)}), 400


if __name__ == "__main__":
    host = os.environ.get("QR_TOOL_HOST", "127.0.0.1")
    port = int(os.environ.get("QR_TOOL_PORT", "5055"))
    app.run(host=host, port=port, debug=False)
