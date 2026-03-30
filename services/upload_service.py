from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

import supabase_client as db
from .match_service import ValidationError

TABLE = "lumi_sn_list"


def _parse_xlsx(file_bytes: bytes) -> list[str]:
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    values: list[str] = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        cell = row[0]
        if cell is None:
            continue
        text = str(cell).strip()
        if text:
            values.append(text)
    wb.close()
    return values


def _parse_csv(file_bytes: bytes) -> list[str]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    values: list[str] = []
    for row in reader:
        if not row:
            continue
        cell = row[0].strip()
        if cell:
            values.append(cell)
    return values


def upload_lumi_sn_file(filename: str, file_bytes: bytes) -> dict[str, Any]:
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xls")):
        sn_list = _parse_xlsx(file_bytes)
    elif lower.endswith(".csv"):
        sn_list = _parse_csv(file_bytes)
    else:
        raise ValidationError("지원하지 않는 파일 형식입니다. xlsx 또는 csv 파일을 업로드해주세요.")

    if not sn_list:
        raise ValidationError("파일에서 SN 데이터를 찾을 수 없습니다.")

    # Deduplicate while preserving order
    seen: set[str] = set()
    unique_list: list[str] = []
    for sn in sn_list:
        if sn not in seen:
            seen.add(sn)
            unique_list.append(sn)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Clear existing list and insert new
    db.delete(TABLE, "id=gt.0")

    # Insert in batches of 100
    batch_size = 100
    for i in range(0, len(unique_list), batch_size):
        batch = unique_list[i : i + batch_size]
        rows = [{"sn_value": sn, "uploaded_at": now} for sn in batch]
        db.insert(TABLE, rows)

    return {
        "total_parsed": len(sn_list),
        "unique_count": len(unique_list),
        "uploaded_at": now,
    }


def get_lumi_sn_list(limit: int = 200) -> list[dict[str, Any]]:
    rows = db.select(
        TABLE,
        order="id.asc",
        limit=limit if limit > 0 else None,
    )
    return [{"id": r["id"], "sn_value": r["sn_value"], "uploaded_at": r["uploaded_at"]} for r in rows]


def get_lumi_sn_count() -> int:
    rows = db.select(TABLE, columns="id")
    return len(rows)


def delete_all_lumi_sn() -> None:
    db.delete(TABLE, "id=gt.0")
